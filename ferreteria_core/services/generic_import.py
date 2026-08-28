import csv
from dataclasses import dataclass,field
from decimal import Decimal,InvalidOperation
from pathlib import Path

from ferreteria_core.money import decimal_a_centavos
from ferreteria_core.repositories import InventoryRepository,ProductRepository
from .products import validar_codigo_barras


@dataclass
class ImportSummary:
    encontrados:int=0;nuevos:int=0;duplicados:int=0;errores:int=0;importados:int=0
    detalles_error:list[str]=field(default_factory=list)


class GenericProductImporter:
    REQUIRED={"codigo_barras","descripcion","precio","existencia"}
    def __init__(self,database):self.database=database
    def analizar(self,path):
        rows=list(_read_rows(path));summary=ImportSummary(encontrados=len(rows));valid=[];seen=set()
        with self.database.connect() as connection:
            for number,row in enumerate(rows,start=2):
                try:
                    item=_validate(row);barcode=item["codigo_barras"]
                    exists=connection.execute("SELECT 1 FROM productos WHERE codigo_barras=?",(barcode,)).fetchone()
                    if barcode in seen or exists:summary.duplicados+=1;continue
                    seen.add(barcode);valid.append(item);summary.nuevos+=1
                except ValueError as exc:summary.errores+=1;summary.detalles_error.append(f"Fila {number}: {exc}")
        return summary,valid
    def importar(self,path):
        summary,valid=self.analizar(path)
        with self.database.transaction() as connection:
            for item in valid:
                values={key:item.get(key) for key in ("codigo_barras","clave","descripcion","marca","categoria","precio_venta","stock_minimo")};values["existencia"]=0
                product=ProductRepository.insert_external(connection,values)
                if item["existencia"]:
                    InventoryRepository.update(connection,product.id,item["existencia"],"AJUSTE",item["existencia"],"IMPORTACION_GENERICA","Existencia inicial importada")
                summary.importados+=1
        return summary


def _read_rows(path):
    path=Path(path);suffix=path.suffix.lower()
    if suffix==".csv":
        with path.open("r",encoding="utf-8-sig",newline="") as stream:yield from csv.DictReader(stream)
    elif suffix==".xlsx":
        from openpyxl import load_workbook
        book=load_workbook(path,read_only=True,data_only=True);sheet=book.active;iterator=sheet.iter_rows(values_only=True)
        try:headers=[str(value or "").strip().lower() for value in next(iterator)]
        except StopIteration:return
        for values in iterator:yield dict(zip(headers,values))
        book.close()
    else:raise ValueError("Formato no compatible; use CSV o XLSX")


def _validate(row):
    normalized={str(k or "").strip().lower():v for k,v in row.items()}
    missing=GenericProductImporter.REQUIRED-set(normalized)
    if missing:raise ValueError(f"Faltan columnas: {', '.join(sorted(missing))}")
    barcode=validar_codigo_barras(str(normalized.get("codigo_barras") or ""));description=str(normalized.get("descripcion") or "").strip()
    if not description:raise ValueError("La descripción es obligatoria")
    try:price=decimal_a_centavos(Decimal(str(normalized.get("precio"))))
    except (InvalidOperation,ValueError) as exc:raise ValueError("Precio inválido") from exc
    existence=_integer(normalized.get("existencia"),"existencia");minimum=_integer(normalized.get("stock_minimo") or 0,"stock_minimo")
    return {"codigo_barras":barcode,"descripcion":description,"precio_venta":price,"existencia":existence,"stock_minimo":minimum,
            "clave":_optional(normalized.get("clave")),"marca":_optional(normalized.get("marca")),"categoria":_optional(normalized.get("categoria"))}


def _integer(value,name):
    try:
        decimal=Decimal(str(value));integer=int(decimal)
    except (InvalidOperation,ValueError):raise ValueError(f"{name} inválida")
    if decimal!=integer or integer<0:raise ValueError(f"{name} debe ser un entero no negativo")
    return integer


def _optional(value):
    value=str(value or "").strip();return value or None
